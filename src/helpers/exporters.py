import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


def generate_excel_report(data: list[dict], course_name: str) -> io.BytesIO:
    """
    Generates an Excel workbook from attendance data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header
    headers = [
        "Student Name",
        "Email",
        "Course",
        "Section",
        "Session Title",
        "Date",
        "Status",
        "Method",
        "Check-in Time",
    ]
    ws.append(headers)

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row in data:
        ws.append([
            row["student_name"],
            row["student_email"],
            f"{row['course_name']} ({row['course_code']})",
            row["section_name"],
            row["session_title"],
            row["session_date"].strftime("%Y-%m-%d %H:%M") if row["session_date"] else "N/A",
            row["status"],
            row["method"],
            row["checkin_time"].strftime("%H:%M:%S") if row["checkin_time"] else "N/A",
        ])

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(data: list[dict], course_name: str) -> io.BytesIO:
    """
    Generates a PDF report from attendance data.
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(f"Attendance Report: {course_name}", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Table Header
    table_data = [[
        "Student Name",
        "Section",
        "Session",
        "Date",
        "Status",
        "Method",
        "Time"
    ]]

    # Data rows
    for row in data:
        table_data.append([
            row["student_name"],
            row["section_name"],
            row["session_title"],
            row["session_date"].strftime("%Y-%m-%d") if row["session_date"] else "N/A",
            row["status"],
            row["method"],
            row["checkin_time"].strftime("%H:%M") if row["checkin_time"] else "N/A",
        ])

    # Create Table
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))

    elements.append(t)
    doc.build(elements)
    output.seek(0)
    return output
